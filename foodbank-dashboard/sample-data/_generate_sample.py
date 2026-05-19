"""
Generate a realistic wide-sparse food-bank intake workbook for the Pantry
Dashboard. Modeled on a JotForm-style wide-sparse intake pattern, where each
department only populates the columns relevant to its operation, leaving the
rest of the row blank on purpose. The shape mirrors real food-bank intake
operations without referencing any specific organization.

Run once. Produces pantry-sample.xlsx in this directory.
"""

import random
from collections import defaultdict
from datetime import date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

random.seed(7)

HEADERS = [
    "Submission ID",
    "Email",
    "Date",
    "Department",
    "Donor Name",
    "Donor Type",
    "Pounds Received",
    "Item Count",
    "Recipient Site",
    "Distributed Pounds",
    "Households Served",
    "Notes",
]

DEPARTMENTS = {
    "Seafood":      {"item_lo": 4,  "item_hi": 40,  "lb_lo": 30,  "lb_hi": 220, "season": [1.0,0.9,1.0,1.1,1.0,1.1,1.0,1.0,1.0,1.0,1.2,1.0]},
    "Dairy":        {"item_lo": 20, "item_hi": 120, "lb_lo": 60,  "lb_hi": 480, "season": [1.0,1.0,1.0,1.0,1.0,0.9,0.85,0.9,1.0,1.0,1.15,1.2]},
    "Meat":         {"item_lo": 12, "item_hi": 80,  "lb_lo": 80,  "lb_hi": 520, "season": [1.0,0.95,0.95,1.0,1.0,0.95,0.9,0.95,1.0,1.05,1.35,1.4]},
    "Frozen":       {"item_lo": 30, "item_hi": 220, "lb_lo": 90,  "lb_hi": 620, "season": [1.0,0.95,0.95,0.95,0.95,0.85,0.8,0.85,1.0,1.05,1.2,1.3]},
    "Produce":      {"item_lo": 40, "item_hi": 320, "lb_lo": 120, "lb_hi": 780, "season": [0.9,0.85,0.9,1.05,1.2,1.3,1.35,1.3,1.2,1.1,1.0,0.95]},
    "Bakery":       {"item_lo": 25, "item_hi": 220, "lb_lo": 40,  "lb_hi": 360, "season": [1.0,0.95,0.95,1.0,1.0,1.0,0.95,1.0,1.0,1.05,1.25,1.3]},
    "Canned Goods": {"item_lo": 50, "item_hi": 400, "lb_lo": 60,  "lb_hi": 540, "season": [1.05,1.0,1.0,1.0,1.0,0.9,0.85,0.95,1.0,1.05,1.2,1.25]},
    "Beverages":    {"item_lo": 30, "item_hi": 260, "lb_lo": 40,  "lb_hi": 320, "season": [1.0,1.0,1.0,1.0,1.05,1.15,1.2,1.15,1.0,1.0,1.05,1.05]},
    "Pizza":        {"item_lo": 6,  "item_hi": 60,  "lb_lo": 15,  "lb_hi": 140, "season": [1.0,1.0,1.0,1.0,1.0,1.0,1.0,1.05,1.05,1.05,1.1,1.1]},
    "Other":        {"item_lo": 10, "item_hi": 180, "lb_lo": 20,  "lb_hi": 280, "season": [1.0]*12},
}

DONORS = {
    "Grocery": [
        "Wegmans Exton", "Giant Food Stores", "ACME Markets Exton",
        "Whole Foods Devon", "Trader Joe's Wayne", "Aldi West Chester",
        "Costco Wholesale", "Kimberton Whole Foods", "Sprouts Glen Mills",
    ],
    "Restaurant": [
        "Domino's Pizza Coatesville", "Panera Bread Exton", "Chick-fil-A Lionville",
        "Wawa Distribution", "Tacconelli's Pizza", "Pat's Pizza West Chester",
        "Mission BBQ Exton",
    ],
    "Farm": [
        "Local Farms Co-op", "Pete's Produce Farm", "Wyebrook Farm",
        "Penn State Extension", "Local Harvest Drive",
    ],
    "Corporate": [
        "Vanguard Group", "QVC Studio Park", "Endo Pharmaceuticals",
        "Teleflex", "Lincoln Financial",
    ],
    "Faith": [
        "St. Agnes Parish", "First Presbyterian Church",
        "Community United Methodist", "Temple Brit Shalom",
    ],
    "Individual": [
        "Individual Donations", "Anonymous Donor", "Boy Scout Troop 24",
        "Daisy Troop 803", "Backyard Garden Drive",
    ],
}

DEPT_DONOR_PREF = {
    "Seafood":      ["Grocery", "Corporate"],
    "Dairy":        ["Grocery", "Farm"],
    "Meat":         ["Grocery", "Farm"],
    "Frozen":       ["Grocery", "Corporate"],
    "Produce":      ["Farm", "Grocery", "Individual"],
    "Bakery":       ["Grocery", "Restaurant"],
    "Canned Goods": ["Faith", "Corporate", "Individual", "Grocery"],
    "Beverages":    ["Corporate", "Grocery"],
    "Pizza":        ["Restaurant"],
    "Other":        ["Faith", "Individual", "Corporate"],
}

SITES = [
    "Coatesville Pantry", "West Chester Site", "Phoenixville Hub",
    "Kennett Square Center", "Oxford Distribution", "Downingtown Pantry",
    "Parkesburg Site", "Mobile Unit (Rotating)", "Spring City Pantry",
]

STAFF_EMAILS = [
    "intake1@example.org", "intake2@example.org", "intake3@example.org",
    "warehouse@example.org", "logistics@example.org", "donations@example.org",
    "operations@example.org", "volunteers@example.org",
]

NOTE_SAMPLES = [
    "", "", "", "", "", "",
    "Pickup arrived 30 min late.",
    "Mixed pallet, sorted on site.",
    "Frozen chain intact on arrival.",
    "Repack required before distribution.",
    "First-time donor.",
    "Holiday drive contribution.",
    "Truck overcap; deferred remainder.",
    "Damaged outer packaging only.",
]


def pick_dept(month):
    weights = []
    depts = list(DEPARTMENTS.keys())
    for d in depts:
        weights.append(DEPARTMENTS[d]["season"][month - 1])
    return random.choices(depts, weights=weights, k=1)[0]


def pick_donor(dept):
    prefs = DEPT_DONOR_PREF[dept]
    dtype = random.choices(prefs, k=1)[0] if random.random() < 0.85 else random.choice(list(DONORS.keys()))
    return random.choice(DONORS[dtype]), dtype


def gen_pounds(dept, month):
    cfg = DEPARTMENTS[dept]
    base = random.uniform(cfg["lb_lo"], cfg["lb_hi"])
    base *= cfg["season"][month - 1]
    base *= random.uniform(0.7, 1.25)
    return round(max(5, base), 1)


def gen_items(dept):
    cfg = DEPARTMENTS[dept]
    return random.randint(cfg["item_lo"], cfg["item_hi"])


def gen_row(d):
    dept = pick_dept(d.month)
    donor, dtype = pick_donor(dept)
    pounds = gen_pounds(dept, d.month)
    items = gen_items(dept)

    role = random.choices(["intake", "distribution", "both"], weights=[65, 25, 10])[0]

    submission_id = str(random.randint(6_000_000_000_000_000_000, 6_400_000_000_000_000_000))
    email = random.choice(STAFF_EMAILS)
    note = random.choice(NOTE_SAMPLES)

    row = {h: None for h in HEADERS}
    row["Submission ID"] = submission_id
    row["Email"] = email
    row["Date"] = d
    row["Department"] = dept
    row["Notes"] = note if note else None

    if role in ("intake", "both"):
        row["Donor Name"] = donor
        row["Donor Type"] = dtype
        row["Pounds Received"] = pounds
        if random.random() < 0.75:
            row["Item Count"] = items

    if role in ("distribution", "both"):
        site = random.choice(SITES)
        dist = round(pounds * random.uniform(0.78, 1.05), 1) if role == "both" else round(gen_pounds(dept, d.month) * 0.9, 1)
        hh = max(1, int(dist / random.uniform(8, 14)))
        row["Recipient Site"] = site
        row["Distributed Pounds"] = dist
        row["Households Served"] = hh

    return row


def main():
    start = date(2025, 5, 1)
    end = date(2026, 4, 30)
    days = (end - start).days

    n = random.randint(360, 460)
    rows = []
    for _ in range(n):
        d = start + timedelta(days=random.randint(0, days))
        rows.append(gen_row(d))

    rows.sort(key=lambda r: r["Date"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Form responses"

    ws.append(HEADERS)
    for r in rows:
        ws.append([r[h] for h in HEADERS])

    header_font = Font(bold=True, color="EDEDEB")
    header_fill = PatternFill(start_color="15151D", end_color="15151D", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    date_col = HEADERS.index("Date") + 1
    for row in ws.iter_rows(min_row=2, min_col=date_col, max_col=date_col):
        for cell in row:
            cell.number_format = "yyyy-mm-dd"

    widths = [22, 24, 12, 16, 28, 14, 16, 14, 24, 18, 18, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws2 = wb.create_sheet("Monthly Rollup")
    ws2.append([
        "Month", "Total Pounds Received", "Total Distributed Pounds",
        "Households Served", "Distinct Donors", "Distinct Departments",
    ])

    monthly = defaultdict(list)
    for r in rows:
        key = r["Date"].strftime("%Y-%m")
        monthly[key].append(r)

    for k in sorted(monthly):
        group = monthly[k]
        recv = round(sum(g["Pounds Received"] or 0 for g in group), 1)
        dist = round(sum(g["Distributed Pounds"] or 0 for g in group), 1)
        hh = sum(g["Households Served"] or 0 for g in group)
        donors = len({g["Donor Name"] for g in group if g["Donor Name"]})
        depts = len({g["Department"] for g in group if g["Department"]})
        ws2.append([k, recv, dist, hh, donors, depts])

    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for i, w in enumerate([12, 24, 26, 20, 18, 22], start=1):
        ws2.column_dimensions[chr(64 + i)].width = w

    out = "/sessions/adoring-busy-volta/mnt/Null Set Labs Website/foodbank-dashboard/sample-data/pantry-sample.xlsx"
    wb.save(out)

    sparsity = {h: sum(1 for r in rows if r[h] is None) for h in HEADERS}
    print(f"Wrote {out}")
    print(f"Rows: {len(rows)}; months: {len(monthly)}")
    print("Empty-cell counts by column (wide-sparse pattern):")
    for h in HEADERS:
        print(f"  {h:24s}  empty: {sparsity[h]:4d} / {len(rows)}")


if __name__ == "__main__":
    main()
in HEADERS:
        print(f"  {h:24s}  empty: {sparsity[h]:4d} / {len(rows)}")


if __name__ == "__main__":
    main()
